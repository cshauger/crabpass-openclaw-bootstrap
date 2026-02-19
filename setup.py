#!/usr/bin/env python3
import os
import json

owner_id = os.environ.get('OWNER_TELEGRAM_ID', '')
model = os.environ.get('OPENCLAW_MODEL', os.environ.get('MODEL', 'groq/llama-3.3-70b-versatile'))
bot_name = os.environ.get('BOT_NAME', 'Assistant')
bot_username = os.environ.get('BOT_USERNAME', '')  # e.g., CrabFresh99Bot

# Config with compaction settings and sandbox allowing exec
config = {
    "gateway": {"mode": "local"},
    "agents": {
        "defaults": {
            "model": {
                "primary": model
            },
            "compaction": {
                "reserveTokensFloor": 4000
            }
        }
    },
    "channels": {
        "telegram": {
            "enabled": True,
            "dmPolicy": os.environ.get('DM_POLICY', 'allowlist'),
        }
    }
}
# Build allowlist from owner + additional users
allowed_users = []
if owner_id:
    allowed_users.append(owner_id)
# Add additional users from ALLOWED_USERS env var (comma-separated)
extra_users = os.environ.get('ALLOWED_USERS', '')
if extra_users:
    for uid in extra_users.split(','):
        uid = uid.strip()
        if uid and uid not in allowed_users:
            allowed_users.append(uid)
if allowed_users:
    config["channels"]["telegram"]["allowFrom"] = allowed_users

config_data = json.dumps(config, indent=2)

# Write config
state_dir = os.environ.get('OPENCLAW_STATE_DIR', '/home/openclaw/.openclaw')
os.makedirs(state_dir, exist_ok=True)
config_path = os.path.join(state_dir, 'config.json5')
with open(config_path, 'w') as f:
    f.write(config_data)
print(f"Config written to {config_path}:")
print(config_data)

# Create workspace
workspace_dir = os.path.join(state_dir, 'workspace')
os.makedirs(workspace_dir, exist_ok=True)

# SOUL.md with email info
email_address = f"{bot_username.lower()}@crabpass.ai" if bot_username else "yourbot@crabpass.ai"
soul_path = os.path.join(workspace_dir, 'SOUL.md')
with open(soul_path, 'w') as f:
    f.write(f"""# {bot_name}

You are a helpful AI assistant powered by CrabPass.

## Your Email
Your email address is: **{email_address}**
People can email you and you can check your inbox.

## Checking Email
To check your emails:
```bash
curl -s "https://email-webhook-production-887d.up.railway.app/emails?bot_username={bot_username or 'YOUR_BOT_USERNAME'}&unread_only=true"
```

## Web Browsing
You can fetch web pages using curl:
```bash
curl -s "https://example.com" | head -100
```

For search, use DuckDuckGo HTML:
```bash
curl -s "https://html.duckduckgo.com/html/?q=your+search+query" | grep -oP '(?<=<a rel="nofollow" class="result__a" href=")[^"]*' | head -5
```

## OneDrive Storage
You have access to the Spex folder on OneDrive via rclone:
```bash
# List files
rclone ls onedrive:Spex/

# Download a file
rclone copy "onedrive:Spex/filename.xlsx" /tmp/

# Upload a file
rclone copy /tmp/myfile.txt onedrive:Spex/
```

## Reading Excel Files
You can read Excel files using Python openpyxl:
```python
import openpyxl
wb = openpyxl.load_workbook('/tmp/file.xlsx', data_only=True)
print(wb.sheetnames)
ws = wb.active
for row in ws.iter_rows(min_row=1, max_row=10, values_only=True):
    print(row)
```

Workflow: download with rclone → read with openpyxl → process data.

## Personality
Be helpful, concise, and friendly. You have access to email, web browsing via curl, OneDrive storage, and other capabilities through CrabPass.
""")
print(f"Created SOUL.md with email info")

# Create email skill
skills_dir = os.path.join(workspace_dir, 'skills', 'email')
os.makedirs(skills_dir, exist_ok=True)
skill_path = os.path.join(skills_dir, 'SKILL.md')
with open(skill_path, 'w') as f:
    f.write(f"""# Email Skill

## Your Email Address
`{email_address}`

## Check Emails
```bash
curl -s "https://email-webhook-production-887d.up.railway.app/emails?bot_username={bot_username or 'YOUR_BOT_USERNAME'}"
```

Add `&unread_only=true` for unread only.

## Mark as Read
```bash
curl -s -X POST "https://email-webhook-production-887d.up.railway.app/emails/EMAIL_ID/read"
```
""")
print(f"Created email skill at {skill_path}")

# Remove default files
for fname in ['AGENTS.md', 'BOOTSTRAP.md', 'USER.md', 'MEMORY.md']:
    fpath = os.path.join(workspace_dir, fname)
    if os.path.exists(fpath):
        os.remove(fpath)
        print(f"Removed {fname}")

# Setup rclone config if provided (base64 encoded)
import base64
rclone_config_b64 = os.environ.get('RCLONE_CONFIG_B64', '')
if rclone_config_b64:
    rclone_dir = os.path.expanduser('~/.config/rclone')
    os.makedirs(rclone_dir, exist_ok=True)
    rclone_path = os.path.join(rclone_dir, 'rclone.conf')
    try:
        rclone_config = base64.b64decode(rclone_config_b64).decode('utf-8')
        with open(rclone_path, 'w') as f:
            f.write(rclone_config)
        print(f"Wrote rclone config to {rclone_path}")
    except Exception as e:
        print(f"Failed to write rclone config: {e}")
